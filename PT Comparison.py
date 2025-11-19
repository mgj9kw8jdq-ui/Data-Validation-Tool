# -*- coding: utf-8 -*-
"""
Created on Fri May 30 09:38:17 2025

@author: 20353120
"""

import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
# Load Excel files
summary_df = pd.read_excel("D:\OneDrive - Larsen & Toubro\Desktop\PT Summary.xlsx", dtype={"Report Number": str})
master_df = pd.read_excel("D:\OneDrive - Larsen & Toubro\Desktop\MASTER LHS PT.xlsx", dtype={"Report Number": str})
# Standardize and clean
summary_df.columns = summary_df.columns.str.strip()
master_df.columns = master_df.columns.str.strip()

results_summary_vs_master = []
results_master_vs_summary = []
# === Summary → Master Comparison ===
for _, row in tqdm(summary_df.iterrows(), total=len(summary_df), desc="Comparing Summary → Master"):
   rep_no = row["Report Number"]
   rep_date = row["Report Date"]
   line = row["Line Number"]
   joint = row["Joint Number"]
   material = row.get("Material", None)
   welder = row.get("Welder", None)
   matches = master_df[master_df["Report Number"] == rep_no]
   if matches.empty:
       status = "Missing in Master"
       remarks = "Report Number not found in Master"
       match_row_str = "-"
   else:
       # OR logic: match line OR joint
       possible_matches = matches[
           (matches["Line Number"] == line) | (matches["Joint Number"] == joint)
       ]
       if not possible_matches.empty:
           found = False
           for _, ref in possible_matches.iterrows():
               line_match = ref["Line Number"] == line
               joint_match = ref["Joint Number"] == joint
               material_match = ref["Material"] == material
               welder_match = ref["Welder"] == welder
               date_match = pd.to_datetime(ref["Report Date"]) == pd.to_datetime(rep_date)
               if line_match and joint_match:
                   if material_match and welder_match and date_match:
                       status = "Match"
                       remarks = "-"
                   else:
                       status = "Mismatch"
                       remarks = (
                           f"Summary: {rep_date}, {material}, {welder} | "
                           f"Master: {ref['Report Date']}, {ref['Material']}, {ref['Welder']}"
                       )
               else:
                   mismatch_fields = []
                   if not line_match:
                       mismatch_fields.append("Line Number")
                   if not joint_match:
                       mismatch_fields.append("Joint Number")
                   status = "Mismatch"
                   remarks = f"{' and '.join(mismatch_fields)} mismatch"
               match_row_str = (
                   f"{ref['Report Date']} | {ref['Line Number']} | "
                   f"{ref['Joint Number']} | {ref['Material']} | {ref['Welder']}"
               )
               found = True
               break
           if not found:
               status = "Mismatch"
               remarks = "Partial match not found"
               match_row_str = "-"
       else:
           status = "Mismatch"
           remarks = "Line and Joint Number do not match under same Report Number"
           match_row_str = "-"
   results_summary_vs_master.append([
       rep_no, rep_date, line, joint, material, welder,
       status, remarks, match_row_str
   ])
# === Master → Summary Comparison ===
for _, row in tqdm(master_df.iterrows(), total=len(master_df), desc="Comparing Master → Summary"):
   rep_no = row["Report Number"]
   rep_date = row["Report Date"]
   line = row["Line Number"]
   joint = row["Joint Number"]
   material = row.get("Material", None)
   welder = row.get("Welder", None)
   matches = summary_df[summary_df["Report Number"] == rep_no]
   if matches.empty:
       status = "Missing in Summary"
       remarks = "Report Number not found in Summary"
       match_row_str = "-"
   else:
       possible_matches = matches[
           (matches["Line Number"] == line) | (matches["Joint Number"] == joint)
       ]
       if not possible_matches.empty:
           found = False
           for _, ref in possible_matches.iterrows():
               line_match = ref["Line Number"] == line
               joint_match = ref["Joint Number"] == joint
               material_match = ref["Material"] == material
               welder_match = ref["Welder"] == welder
               date_match = pd.to_datetime(ref["Report Date"]) == pd.to_datetime(rep_date)
               if line_match and joint_match:
                   if material_match and welder_match and date_match:
                       status = "Match"
                       remarks = "-"
                   else:
                       status = "Mismatch"
                       remarks = (
                           f"Master: {rep_date}, {material}, {welder} | "
                           f"Summary: {ref['Report Date']}, {ref['Material']}, {ref['Welder']}"
                       )
               else:
                   mismatch_fields = []
                   if not line_match:
                       mismatch_fields.append("Line Number")
                   if not joint_match:
                       mismatch_fields.append("Joint Number")
                   status = "Mismatch"
                   remarks = f"{' and '.join(mismatch_fields)} mismatch"
               match_row_str = (
                   f"{ref['Report Date']} | {ref['Line Number']} | "
                   f"{ref['Joint Number']} | {ref['Material']} | {ref['Welder']}"
               )
               found = True
               break
           if not found:
               status = "Mismatch"
               remarks = "Partial match not found"
               match_row_str = "-"
       else:
           status = "Mismatch"
           remarks = "Line and Joint Number do not match under same Report Number"
           match_row_str = "-"
   results_master_vs_summary.append([
       rep_no, rep_date, line, joint, material, welder,
       status, remarks, match_row_str
   ])

# === Save the Results ===
summary_vs_master_df = pd.DataFrame(results_summary_vs_master, columns=[
   "Report Number", "Report Date", "Line Number", "Joint Number", "Material", "Welder",
   "Comparison Status", "Remarks", "Matched Master Row"
])
master_vs_summary_df = pd.DataFrame(results_master_vs_summary, columns=[
   "Report Number", "Report Date", "Line Number", "Joint Number", "Material", "Welder",
   "Comparison Status", "Remarks", "Matched Summary Row"
])

excel_path = "D:\OneDrive - Larsen & Toubro\Desktop\Comparison_Result Final.xlsx"
with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
   summary_vs_master_df.to_excel(writer, sheet_name="Summary_vs_Master", index=False)
   master_vs_summary_df.to_excel(writer, sheet_name="Master_vs_Summary", index=False)

# Add color formatting
wb = load_workbook(excel_path)
fill_match = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")    # green
fill_mismatch = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # orange
fill_missing = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")   # red
for sheet_name in ["Summary_vs_Master", "Master_vs_Summary"]:
   ws = wb[sheet_name]
   for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
       status = row[6].value  # 7th column: Comparison Status
       if status == "Match":
           fill = fill_match
       elif "Missing" in status:
           fill = fill_missing
       else:
           fill = fill_mismatch
       for cell in row:
           cell.fill = fill
wb.save(excel_path)
print("✅ Comparison completed with color coding. File saved as 'PT_Comparison_Result.xlsx'")