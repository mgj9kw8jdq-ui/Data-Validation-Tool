# -*- coding: utf-8 -*-
"""
Created on Fri May 30 09:38:17 2025

@author: 20353120
"""

import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
# Load Excel files
summary_df = pd.read_excel("D:\OneDrive - Larsen & Toubro\Desktop\RT Summary.xlsx", dtype={"Report Number": str})
master_df = pd.read_excel("D:\OneDrive - Larsen & Toubro\Desktop\MASTER LHS RT.xlsx", dtype={"Report Number": str})
# Standardize and clean
summary_df.columns = summary_df.columns.str.strip()
master_df.columns = master_df.columns.str.strip()
expected_columns = ["Report Number", "Report Date", "Line Number", "Joint Number", "Material", "Welder"]
summary_df = summary_df[expected_columns]
master_df = master_df[expected_columns]
summary_df["Report Number"] = summary_df["Report Number"].astype(str).str.strip()
master_df["Report Number"] = master_df["Report Number"].astype(str).str.strip()
# Compare and store results
results = []
# Compare Summary against Master
for _, row in tqdm(summary_df.iterrows(), total=len(summary_df), desc="Summary -> Master"):
   rep_no, rep_date, line, joint, material, welder = row
   matches = master_df[master_df["Report Number"] == rep_no]
   same_joint = matches[
       (matches["Line Number"] == line) & (matches["Joint Number"] == joint)
   ]
   
   if same_joint.empty:
       status = "Missing in Master"
       remarks = "Line/Joint not found"
       match_row_str = "-"
   else:
       ref = same_joint.iloc[0]
       mismatches = []
       if pd.to_datetime(ref["Report Date"]) != pd.to_datetime(rep_date):
           mismatches.append(f"Report Date ({rep_date} ≠ {ref['Report Date']})")
       if ref["Material"] != material:
           mismatches.append(f"Material ({material} ≠ {ref['Material']})")
       if ref["Welder"] != welder:
           mismatches.append(f"Welder ({welder} ≠ {ref['Welder']})")
       if not mismatches:
           status = "Match"
           remarks = "-"
       else:
           status = "Mismatch"
           remarks = "Mismatch in: " + ", ".join(mismatches)
       
       match_row_str = (
           f"{ref['Report Date']} | {ref['Line Number']} | {ref['Joint Number']} | "
           f"{ref['Material']} | {ref['Welder']}"
       )
       
   results.append({
       "Source": "Summary",
       "Report Number": rep_no,
       "Report Date": rep_date,
       "Line Number": line,
       "Joint Number": joint,
       "Material": material,
       "Welder": welder,
       "Status": status,
       "Remarks": remarks,
       "Matching Row in Master": match_row_str
   })
# Compare Master against Summary
for _, row in tqdm(master_df.iterrows(), total=len(master_df), desc="Master -> Summary"):
   rep_no, rep_date, line, joint, material, welder = row
   matches = summary_df[summary_df["Report Number"] == rep_no]
   same_joint = matches[
       (matches["Line Number"] == line) & (matches["Joint Number"] == joint)
   ]
   if same_joint.empty:
       results.append({
           "Source": "Master",
           "Report Number": rep_no,
           "Report Date": rep_date,
           "Line Number": line,
           "Joint Number": joint,
           "Material": material,
           "Welder": welder,
           "Status": "Missing in Summary",
           "Remarks": "Line/Joint not found",
           "Matching Row in Master": "-"
       })
# Save results to Excel
result_df = pd.DataFrame(results)
output_file = "D:\OneDrive - Larsen & Toubro\Desktop\Comparison_RT.xlsx"
result_df.to_excel(output_file, index=False)
# Add color formatting
wb = load_workbook(output_file)
ws = wb.active
# Find the 'Status' column index
status_col = None
for i, cell in enumerate(ws[1]):
   if cell.value == "Status":
       status_col = i + 1
       break
# Define fills
fills = {
   "Match": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),       # Green
   "Mismatch": PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid"),   # Yellow
   "Missing in Master": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),  # Red
   "Missing in Summary": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")  # Red
}
# Apply fills
for row in ws.iter_rows(min_row=2, min_col=status_col, max_col=status_col):
   status = row[0].value
   fill = fills.get(status)
   if fill:
       for cell in row:
           cell.fill = fill
# Save workbook
wb.save(output_file)
print(f"🎉 Comparison completed! Open '{output_file}' to view results.")